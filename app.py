from flask import Flask, render_template, request, redirect
import openpyxl
import os
from werkzeug.utils import secure_filename
from random import randint


app = Flask(__name__, template_folder='templates')



# Hàm để đọc dữ liệu từ sheet "categorty" trong tệp Excel và trả về dưới dạng danh sách
def read_categortys():
    try:
        # Lấy đường dẫn thư mục chứa tệp Excel
        directory = 'data'
        
        # Tạo đường dẫn đầy đủ tới tệp Excel
        file_path = os.path.join(directory, 'data.xlsx')

        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook['categorty']  # Chọn sheet có tên là "categorty"
        categortys = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            categorty = {
                'id': row[0],
                'name': row[1]
            }
            categortys.append(categorty)

        workbook.close()
        return categortys
    except FileNotFoundError:
        return []

def read_products():
    try:
        # Lấy đường dẫn thư mục chứa tệp Excel
        directory = 'data'
        
        # Tạo đường dẫn đầy đủ tới tệp Excel
        file_path = os.path.join(directory, 'data.xlsx')

        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook['products']
        products = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            product = {
                'id': row[0],
                'image': row[1],
                'key': row[2],
                'name': row[3],
                'categorty': row[4],
                'unit': row[5],
                'quantity': row[6]

            }
            products.append(product)

        workbook.close()
        return products
    except FileNotFoundError:
        return []
# Hàm để ghi dữ liệu vào sheet "categorty" trong tệp Excel
def write_categortys(categortys):
    try:
        # Lấy đường dẫn thư mục chứa tệp Excel
        directory = 'data'
        
        # Tạo đường dẫn đầy đủ tới tệp Excel
        file_path = os.path.join(directory, 'data.xlsx')
        
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.create_sheet(title='categorty')  # Tạo sheet mới có tên là "categorty"

    worksheet = workbook['categorty']  # Lấy hoặc tạo sheet "categorty"
    worksheet.delete_rows(2, worksheet.max_row)  # Xóa tất cả dòng trừ dòng đầu tiên (dòng tiêu đề)

    for categorty in categortys:
        worksheet.append([categorty['id'], categorty['name']])

    workbook.save(file_path)

def write_products(products):
    try:
        # Lấy đường dẫn thư mục chứa tệp Excel
        directory = 'data'
        
        # Tạo đường dẫn đầy đủ tới tệp Excel
        file_path = os.path.join(directory, 'data.xlsx')
        
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.create_sheet(title='products')

    worksheet = workbook['products']
    worksheet.delete_rows(2, worksheet.max_row)

    for product in products:
        worksheet.append([product['id'], product['image'], product['key'],
                          product['name'], product['categorty'], product['unit']
                          , product['quantity']])

    workbook.save(file_path)
    
@app.route('/')
def home():
    categortys = read_categortys()
    products = read_products()
    return render_template('examples/index.html', categortys=categortys, products=products)


#--Product--
@app.route('/product')
def product():
    categortys = read_categortys()
    return render_template('examples/product.html', categortys=categortys)

@app.route('/add_product', methods=['POST'])
def add_product():
         
    image = request.files['product_image']
    key = request.form['product_key']
    name = request.form['product_name']
    categorty = request.form['product_cate']
    unit = request.form['product_unit']
    quantity = request.form['product_quantity']
    # Lấy tên gốc của tệp hình ảnh
    get_name_image = secure_filename(image.filename)
    name_image, extension = os.path.splitext(get_name_image)
            
    # Tạo tên mới cho tệp hình ảnh dựa trên tên gốc và một số ngẫu nhiên
    new_image = f"{name_image}{str(randint(0, 999)).zfill(3)}{extension}"
    
    # Lưu tệp hình ảnh vào thư mục cụ thể (ví dụ: 'public/uploads/product')
    upload_folder = 'static/assets/img/product'
    file_path = os.path.join(upload_folder, new_image)
    image.save(file_path)

    products = read_products()
    if not products:
        id = 1
    else:
        id = products[-1]['id'] + 1

    product = {
        'id': id,
        'image': new_image,
        'key': key,
        'name': name,
        'categorty': categorty,
        'unit': unit,
        'quantity': quantity
    }

    products.append(product)
    write_products(products)
    # return categorty
    return redirect('/product')

@app.route('/edit_product', methods=['POST'])
def edit_product():

    id = int(request.form['edit_id'])
    key = request.form['edit_key']
    name = request.form['edit_name']
    categorty = request.form['edit_cate']
    unit = request.form['edit_unit']
    quantity = request.form['edit_quantity']

    old_name = request.form['name_old']
    if(request.files['edit_image_name']):
        new_name = request.files['edit_image_name']
        # Lấy tên gốc của tệp hình ảnh
        get_name_image = secure_filename(new_name.filename)
        name_image, extension = os.path.splitext(get_name_image)
                
        # Tạo tên mới cho tệp hình ảnh dựa trên tên gốc và một số ngẫu nhiên
        new_image = f"{name_image}{str(randint(0, 999)).zfill(3)}{extension}"
        
        # Lưu tệp hình ảnh vào thư mục cụ thể (ví dụ: 'public/uploads/product')
        upload_folder = 'static/assets/img/product'
        file_path = os.path.join(upload_folder, new_image)
        old_file_path = os.path.join(upload_folder, old_name)
        os.remove(old_file_path)
        new_name.save(file_path)
        
        try:
            directory = 'data'
            file_path = os.path.join(directory, 'data.xlsx')
            
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            raise Exception("Tệp Excel không tồn tại!")

        worksheet = workbook['products']

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == id:
                worksheet.cell(row=row_number, column=2, value=str(new_image))
                worksheet.cell(row=row_number, column=3, value=str(key))
                worksheet.cell(row=row_number, column=4, value=str(name))
                worksheet.cell(row=row_number, column=5, value=str(categorty))
                worksheet.cell(row=row_number, column=6, value=str(unit))
                worksheet.cell(row=row_number, column=7, value=str(quantity))
                break

        workbook.save(file_path)
        return redirect('/')
    
    else:
        try:
            directory = 'data'
            file_path = os.path.join(directory, 'data.xlsx')
            
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            raise Exception("Tệp Excel không tồn tại!")

        worksheet = workbook['products']

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == id:
                worksheet.cell(row=row_number, column=3, value=str(key))
                worksheet.cell(row=row_number, column=4, value=str(name))
                worksheet.cell(row=row_number, column=5, value=str(categorty))
                worksheet.cell(row=row_number, column=6, value=str(unit))
                worksheet.cell(row=row_number, column=7, value=str(quantity))
                break

        workbook.save(file_path)
        return redirect('/')


@app.route('/delete_product/<int:product_id>')
def delete_product(product_id):
    products = read_products()
    for product in products:
        if product['id'] == product_id:
            products.remove(product)
            write_products(products)
            break

    return redirect('/')

# --Category--
@app.route('/categorty')
def categorty():
    categortys = read_categortys()
    return render_template('examples/categorty.html', categortys=categortys)

@app.route('/add_categorty', methods=['POST'])
def add_categorty():
    name = request.form['categorty_name']
    categortys = read_categortys()
    if not categortys:
        id = 1
    else:
        id = categortys[-1]['id'] + 1

    categorty = {
        'id': id,
        'name': name
    }

    categortys.append(categorty)
    write_categortys(categortys)
    return redirect('/categorty')
    # success_message = f"Meo mei đã thêm '{name}' ❤️"
    # return name
    # return redirect('/categorty')



@app.route('/edit_categorty', methods=['POST'])
def edit_categorty():
    category_id = int(request.form['category_id'])
    old_name = ""
    updated_name = request.form['categorty_rename']
    try:
        directory = 'data'
        file_path = os.path.join(directory, 'data.xlsx')
        
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise Exception("Tệp Excel không tồn tại!")

    worksheet = workbook['categorty']

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == category_id:
            old_name = worksheet.cell(row=row_number, column=2).value
            worksheet.cell(row=row_number, column=2, value=str(updated_name))
            break

    workbook.save(file_path)
    workbook.close()

    worksheet = workbook['products']

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[4] == old_name:
            worksheet.cell(row=row_number, column=5, value=str(updated_name))

    workbook.save(file_path)
    workbook.close()


    return redirect('/categorty')

@app.route('/delete_categorty/<int:categorty_id>')
def delete_categorty(categorty_id):
    categortys = read_categortys()
    for categorty in categortys:
        if categorty['id'] == categorty_id:
            categortys.remove(categorty)
            write_categortys(categortys)
            break

    return redirect('/categorty')

if __name__ == '__main__':
    app.run(debug=True)
